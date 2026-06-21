# -*- coding: utf-8 -*-
"""将 test_assets.json 转换为格式化的 Excel 工作簿。

用法:
    python generate_excel.py <json_path> [-o <output_path>]

    json_path   必需，指向 test_assets.json
    -o, --output  可选，输出 xlsx 路径；默认输出到 json 同目录下的 功能测试用例设计.xlsx
                  若输出文件已存在，自动追加 _1, _2 等后缀。

示例:
    python generate_excel.py test_assets.json
    python generate_excel.py test_assets.json -o ../最终用例/功能测试用例设计.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="微软雅黑", bold=True, size=11)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
LABEL_FONT = Font(name="微软雅黑", bold=True, size=11)
DATA_FONT = Font(name="微软雅黑", size=10)
P0_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
P1_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
MUST_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
ALT_FILL = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def cell(ws, row, col, value, wrap=True, font=None, fill=None):
    """Set a cell value with default styling (border + data font)."""
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = WRAP if wrap else Alignment(vertical="top")
    c.border = BORDER
    c.font = font if font else DATA_FONT
    if fill:
        c.fill = fill
    return c


def border_empty(ws, row, col_start, col_end):
    """Apply border-only styling to empty cells, so merged regions don't leak."""
    for ci in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=ci)
        if c.value is None:
            c.border = BORDER
            c.font = DATA_FONT


def style_header(ws, headers, freeze_row=1):
    """Apply header styling: deep-blue fill, white bold font, border, freeze."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=freeze_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.freeze_panes = f"A{freeze_row + 1}"


def auto_width(ws, min_w=8, max_w=55):
    """Auto-fit column widths, treating CJK chars as ~2 wide."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = 0
        for c in col_cells:
            if not c.value:
                continue
            for line in str(c.value).split("\n"):
                w = sum(2 if ord(ch) > 0x2000 else 1 for ch in line)
                best = max(best, w)
        ws.column_dimensions[col_letter].width = min(max(best + 2, min_w), max_w)


def auto_row_height(ws, min_height=15, extra_per_line=4):
    """Set row heights based on max lines-per-cell in each row.

    Only processes rows where at least one cell has wrap_text enabled.
    The extra_per_line accounts for font height (~12pt line + padding).
    """
    # Determine max column width for the sheet (approximate chars-per-line)
    col_widths = {}
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        dim = ws.column_dimensions.get(col_letter)
        col_widths[col_cells[0].column] = dim.width if dim and dim.width else 12

    for row in ws.iter_rows():
        row_num = row[0].row
        max_lines = 1
        for c in row:
            if not c.value or not c.alignment.wrap_text:
                continue
            col_w = col_widths.get(c.column, 30)
            # Approximate: each CJK char ~2 units, ASCII ~1 unit
            for line in str(c.value).split("\n"):
                char_width = sum(2 if ord(ch) > 0x2000 else 1 for ch in line)
                lines_needed = max(1, -(-char_width // max(1, int(col_w))))
                max_lines = max(max_lines, lines_needed)
            # Also count explicit line breaks
            explicit_lines = str(c.value).count("\n") + 1
            max_lines = max(max_lines, explicit_lines)

        ws.row_dimensions[row_num].height = max(min_height, max_lines * 15 + extra_per_line)


def resolve_output_path(output_dir, base_name="功能测试用例设计.xlsx"):
    """Return a non-conflicting output path, appending _N if needed."""
    target = output_dir / base_name
    if not target.exists():
        return target
    stem = Path(base_name).stem
    suffix = Path(base_name).suffix
    n = 1
    while True:
        alt = output_dir / f"{stem}_{n}{suffix}"
        if not alt.exists():
            return alt
        n += 1


# ── Sheet builders ────────────────────────────────────────

def build_stats_sheet(wb, data):
    ws = wb.create_sheet("统计汇总")
    ws.sheet_properties.tabColor = "1F4E79"

    stats = data.get("stats", {})
    meta = data.get("meta", {})

    # ── helpers ────────────────────────────────────────

    def section_title(text, cols):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        for ci in range(1, cols + 1):
            ws.cell(row=r, column=ci).fill = SECTION_FILL
            ws.cell(row=r, column=ci).border = BORDER
            ws.cell(row=r, column=ci).font = DATA_FONT
        cell(ws, r, 1, text, wrap=False, font=SECTION_FONT, fill=SECTION_FILL)
        r += 1

    def sub_header(cols, headers):
        """Thin-layered header row under a section title."""
        nonlocal r
        for ci, h in enumerate(headers, 1):
            c = cell(ws, r, ci, h, wrap=False)
            c.font = Font(name="微软雅黑", bold=True, size=10)
            c.fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
        for ci in range(len(headers) + 1, cols + 1):
            c = cell(ws, r, ci, "", wrap=False)
            c.fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
        r += 1

    def stats_table(title, cols, headers, rows, highlight_val=None, highlight_fill=None):
        nonlocal r
        section_title(title, cols)
        sub_header(cols, headers)

        total = 0
        for row_vals in rows:
            for ci, v in enumerate(row_vals, 1):
                cell(ws, r, ci, v, wrap=False)
            # Fill any remaining columns with bordered empty cells for uniformity
            for ci in range(len(row_vals) + 1, cols + 1):
                cell(ws, r, ci, "", wrap=False)
            if highlight_val and row_vals and row_vals[0] == highlight_val and highlight_fill:
                for ci in range(1, cols + 1):
                    ws.cell(row=r, column=ci).fill = highlight_fill
            if len(row_vals) > 1 and isinstance(row_vals[1], (int, float)):
                total += row_vals[1]
            r += 1

        # total row
        for ci in range(1, cols + 1):
            c = cell(ws, r, ci, "合计" if ci == 1 else (total if ci == 2 else ""),
                     wrap=False, fill=TOTAL_FILL)
            c.font = Font(name="微软雅黑", bold=True, size=10)
        r += 2

    def label_rows(rows, cols):
        """Key-value rows: A=label (bold), B=value. Value may span to cols."""
        nonlocal r
        for k, v in rows:
            cell(ws, r, 1, k, wrap=False, font=LABEL_FONT)
            if cols > 2:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=cols)
            cell(ws, r, 2, v, wrap=False)
            r += 1

    # ── Layout ─────────────────────────────────────────

    COLS = 3  # max columns used by any section on this sheet

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
    c = ws.cell(row=1, column=1, value="功能测试用例设计 — 统计汇总")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    for ci in range(1, COLS + 1):
        ws.cell(row=1, column=ci).border = BORDER
        ws.cell(row=1, column=ci).font = DATA_FONT

    r = 3

    # Info
    info_rows = [
        ("生成日期", meta.get("generated_at", "")),
        ("评审日期", meta.get("review_date", "")),
        ("评审范围", "、".join(meta.get("modules", []))),
    ]
    label_rows(info_rows, COLS)
    r += 1

    # Module counts  (3 cols)
    module_rows = [(m["module"], m["count"], f"{m['count'] / stats['review_summary']['total_cases'] * 100:.1f}%")
                   for m in stats.get("module_counts", [])]
    stats_table("各模块用例数量", 3, ["模块", "用例数量", "占比"], module_rows)

    # Priority counts  (3 cols)
    total = stats["review_summary"]["total_cases"]
    pri_rows = [(p["priority"], p["count"], f"{p['count'] / total * 100:.1f}%")
                for p in stats.get("priority_counts", [])]
    stats_table("各优先级用例数量", 3, ["优先级", "数量", "占比"], pri_rows,
                highlight_val="P0", highlight_fill=P0_FILL)

    # Type counts  (3 cols)
    type_rows = [(t["type"], t["count"], f"{t['count'] / total * 100:.1f}%")
                 for t in stats.get("type_counts", [])]
    stats_table("各用例类型数量", 3, ["用例类型", "数量", "占比"], type_rows)

    # Coverage  (3 cols)
    section_title("覆盖状态分布", 3)
    sub_header(3, ["覆盖项", "状态", "说明"])
    for cov in stats.get("coverage", []):
        cell(ws, r, 1, cov["dimension"], wrap=False)
        cell(ws, r, 2, cov["status"], wrap=False)
        cell(ws, r, 3, cov["note"])
        r += 1
    r += 1

    # Pending questions by module  (2 cols)
    pq_data = stats.get("pending_questions_by_module", {})
    pq_rows = [(mod, cnt) for mod, cnt in pq_data.items()]
    stats_table("待确认问题数量", COLS, ["所属模块", "问题数量"], pq_rows)

    # Review issues by severity  (2 cols data, 3 cols layout)
    ri_data = stats.get("review_issues_by_severity", {})
    ri_rows = [(sev, cnt) for sev, cnt in ri_data.items()]
    stats_table("评审问题数量（按严重级别）", COLS, ["严重级别", "数量"], ri_rows,
                highlight_val="必须修改", highlight_fill=MUST_FILL)

    # Review overview  (2 cols)
    section_title("评审概览", COLS)
    overview = stats.get("review_summary", {})
    ov_rows = [
        ("评审用例总数", str(overview.get("total_cases", ""))),
        ("通过", str(overview.get("passed", ""))),
        ("需修改", str(overview.get("need_modify", ""))),
        ("需确认", str(overview.get("need_confirm", ""))),
        ("涉及模块", overview.get("modules", "")),
        ("评审结论", overview.get("conclusion", "")),
    ]
    label_rows(ov_rows, COLS)

    auto_width(ws)
    auto_row_height(ws)


def build_cases_sheet(wb, data):
    ws = wb.create_sheet("测试用例")

    headers = [
        "用例ID", "模块", "子模块", "需求来源", "测试点ID", "用例标题",
        "优先级", "用例类型", "前置条件", "测试数据", "操作步骤", "预期结果",
        "清理动作", "备注"
    ]
    style_header(ws, headers)

    cases = data.get("test_cases", [])
    field_map = [
        ("id", False), ("module", False), ("sub_module", False),
        ("requirement_source", True), ("test_point_id", False),
        ("title", True), ("priority", False), ("type", False),
        ("precondition", True), ("test_data", True),
        ("steps", True), ("expected", True),
        ("cleanup", True), ("notes", True),
    ]

    for ri, case in enumerate(cases):
        row = ri + 2
        priority = case.get("priority", "")
        is_p0 = priority == "P0"
        is_p1 = priority == "P1"
        is_odd = ri % 2 == 0

        for ci, (field, _) in enumerate(field_map):
            val = case.get(field, "")
            c = cell(ws, row, ci + 1, val, wrap=True)
            if is_p0:
                c.fill = P0_FILL
            elif is_p1:
                c.fill = P1_FILL
            elif is_odd:
                c.fill = ALT_FILL

    ws.auto_filter.ref = f"A1:N{len(cases) + 1}"
    auto_width(ws)
    auto_row_height(ws)


def build_questions_sheet(wb, data):
    ws = wb.create_sheet("待确认问题")

    headers = ["问题编号", "所属模块", "需求位置", "问题描述", "对测试设计的影响", "建议确认对象", "当前处理方式"]
    style_header(ws, headers)

    questions = data.get("pending_questions", [])
    fields = ["id", "module", "requirement_location", "description", "impact", "confirm_with", "handling"]

    for ri, q in enumerate(questions):
        row = ri + 2
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, f in enumerate(fields):
            cell(ws, row, ci + 1, q.get(f, ""), fill=fill)

    ws.auto_filter.ref = f"A1:G{len(questions) + 1}"
    auto_width(ws)
    auto_row_height(ws)


def build_review_sheet(wb, data):
    ws = wb.create_sheet("评审问题")

    headers = ["问题编号", "严重级别", "问题类型", "涉及用例", "涉及需求", "问题描述", "修改建议", "状态"]
    style_header(ws, headers)

    issues = data.get("review_issues", [])
    fields = ["id", "severity", "type", "related_cases", "related_requirement", "description", "suggestion", "status"]

    for ri, issue in enumerate(issues):
        row = ri + 2
        is_must = issue.get("severity", "") == "必须修改"
        is_odd = ri % 2 == 0
        for ci, f in enumerate(fields):
            c = cell(ws, row, ci + 1, issue.get(f, ""))
            if is_must:
                c.fill = MUST_FILL
            elif is_odd:
                c.fill = ALT_FILL

    ws.auto_filter.ref = f"A1:H{len(issues) + 1}"
    auto_width(ws)
    auto_row_height(ws)


# ── Main ──────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="将 test_assets.json 转换为格式化的 Excel 工作簿"
    )
    parser.add_argument(
        "json_path",
        help="JSON 数据文件路径（必需）",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="输出 xlsx 路径。默认输出到 JSON 同目录下的 功能测试用例设计.xlsx",
    )
    args = parser.parse_args()

    json_path = Path(args.json_path)
    if not json_path.exists():
        print(f"ERROR: JSON file not found: {json_path}")
        sys.exit(1)

    # Determine output path
    if args.output:
        output_dir = Path(args.output).parent
        output_name = Path(args.output).name
    else:
        output_dir = json_path.parent
        output_name = "功能测试用例设计.xlsx"

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = resolve_output_path(output_dir, output_name)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    build_stats_sheet(wb, data)
    build_questions_sheet(wb, data)
    build_review_sheet(wb, data)
    build_cases_sheet(wb, data)

    # Remove default empty sheet created by Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.move_sheet("统计汇总", offset=-3)

    wb.save(str(output_path))

    print(f"Excel saved: {output_path}")
    print(f"  Sheets: 统计汇总 | 测试用例 ({len(data.get('test_cases', []))} items) | "
          f"待确认问题 ({len(data.get('pending_questions', []))} items) | "
          f"评审问题 ({len(data.get('review_issues', []))} items)")


if __name__ == "__main__":
    main()
